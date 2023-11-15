import openpyxl
import re
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl import load_workbook
from copy import copy

def append_columns(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Добавляем новые заголовки в конце рабочего листа
    worksheet.cell(row=1, column=worksheet.max_column + 1, value="Сб/комп")
    worksheet.cell(row=1, column=worksheet.max_column + 1, value="Тип объекта")
    worksheet.cell(row=1, column=worksheet.max_column + 1, value="откуда")

    workbook.save(file_path)

    return file_path

def ps10(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Создаем пустой словарь для хранения уникальных элементов
    word_dict = {}

    # Проходим по каждой строке (начиная с второй строки)
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=4).value  # Столбец D

        # Проверяем, есть ли слово в списке и нет ли его уже в word_dict
        if cell_value is not None and cell_value not in word_dict:
            word_dict[cell_value] = True  # Используем True для обозначения наличия элемента

    # Используем словарь для проверки и записи значений в столбец H
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=6).value  # Столбец F

        # Проверяем, есть ли слово в словаре word_dict
        if cell_value in word_dict:
            worksheet.cell(row=row, column=8, value="Cборка")  # Столбец H
        else:
            worksheet.cell(row=row, column=8, value="Компоненты")  # Столбец H

    # Создаем словари Крепеж и Декор
    krepezh_set = {"Болт", "Гайка", "Штифт", "Шайба", "Шуруп"}
    dekor_set = {"Коврики", "Подушки"}

    # Проходим по каждой строке (начиная с второй строки)
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=6).value  # Столбец F

        # Проверяем, к какой категории относится строка
        if cell_value in word_dict:
            category = "Сборка"
        elif cell_value in krepezh_set:
            category = "Крепеж"
        elif cell_value in dekor_set:
            category = "Декор"
        else:
            category = "Детали"

        # Записываем категорию в столбец I
        worksheet.cell(row=row, column=9, value=category)  # Столбец I

    # Проходим по каждой строке (начиная с второй строки)
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=5).value  # Столбец Е

        # метод isascii() проверяет, содержатся ли все символы в строке cell_value в стандарте ASCII (American Standard Code for Information Interchange).
        # Стандарт ASCII включает в себя латинские буквы (как заглавные, так и строчные) и цифры, а также некоторые специальные символы.
        if cell_value.isascii():
            category = "Иностранное"  # Содержит только латиницу и цифры
        elif "ГОСТ" in cell_value:
            category = "РФ по ГОСТ"  # Содержит слово "ГОСТ"
        else:
            category = "РФ"  # Содержит кириллические символы, кроме "ГОСТ"

        # Записываем категорию в столбец J
        worksheet.cell(row=row, column=10, value=category)  # Столбец J

    # Сохраняем изменения в файл
    workbook.save(file_path)
def check_headers(file_path):
    expected_headers = ['кодизделия', 'наименованиеизделия', 'кодсборки', 'наименованиесборки', 'кодкомпоненты', 'наименованиекомпоненты', 'кол-вонасборку']
    df = pd.read_excel(file_path, header=0)

    # Получаем строки с заголовками
    header_row_values = df.columns

    # Преобразуем строки заголовков к формату ожидаемых заголовков (удаляем пробелы и приводим к нижнему регистру)
    actual_headers = [re.sub(r'\s', '', str(header)).lower() for header in header_row_values]

    if actual_headers != expected_headers:
        print(f"Ошибка: Структура заголовков в файле '{file_path}' не соответствует ожидаемой.")
        print(actual_headers)
        print(expected_headers)
        return False
    return True



#функция по нахождению уникальных и повторяющихся строк
def compare_files(df1, df2):
    # Нахождение одинаковых строк
    common_rows = pd.merge(df1, df2, how='inner')

    # Нахождение уникальных строк в первой таблице
    unique_rows_df1 = pd.concat([df1, common_rows]).drop_duplicates(keep=False)

    # Нахождение уникальных строк во второй таблице
    unique_rows_df2 = pd.concat([df2, common_rows]).drop_duplicates(keep=False)

    return common_rows, unique_rows_df1, unique_rows_df2

#функция по редактированию стиля итогового файла
def copy_column_style(ws, source_column_letter, target_column_letters):
    for row in ws.iter_rows(min_col=ws[source_column_letter][0].column,
                            max_col=ws[source_column_letter][0].column,
                            min_row=ws.min_row,
                            max_row=ws.max_row):

        source_cell = row[0] # так как мы имеем дело только с одним столбцом
        source_style = copy(source_cell.style)

        for target_column_letter in target_column_letters:
            target_cell = ws[target_column_letter + str(source_cell.row)]
            copy_style = source_cell._style
            target_cell._style = copy_style





if __name__ == "__main__":
    file1_name = input("Введите название первого файла: ")
    file2_name = input("Введите название второго файла: ")
    # file1_name = 'Машина 1.xlsx'
    # file2_name = 'Машина 2.xlsx'
    # Загрузка данных из файлов Excel
    df1 = pd.read_excel(file1_name)
    df2 = pd.read_excel(file2_name)
    if check_headers(file1_name) and check_headers(file2_name):
        common_rows, unique_rows_df1, unique_rows_df2 = compare_files(df1, df2)
        # Вывод результатов
        print("Общие строки:")
        print(common_rows)

        print("\nУникальные строки в первой таблице:")
        print(unique_rows_df1)

        print("\nУникальные строки во второй таблице:")
        print(unique_rows_df2)
    else:
        print("Ошибка: Структура заголовков в одном из файлов не соответствует ожидаемой.")

    # Загрузка второго файла Excel
    file_path = file2_name
    wb = load_workbook(file_path)

    # Получение активного листа
    sheet = wb.active

    # Загрузка данных из unique_rows_df1
    data_to_append = unique_rows_df1.values.tolist()

    # Определение столбцов второго файла Excel
    columns = list(unique_rows_df1.columns)

    # Определение номера строки, начиная с которой нужно вставлять данные
    start_row = sheet.max_row + 1

    # Сохранение стиля для последующего использования
    reference_cell = sheet.cell(row=2, column=1)
    reference_style = reference_cell._style

    # Вставка данных в конец файла
    for row_data in data_to_append:
        sheet.append(row_data)

    # Применение сохраненного стиля к добавленным строкам (кроме цвета)
    for row_num in range(start_row, sheet.max_row + 1):
        for col_num in range(1, len(columns) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell._style = reference_style

    # Установка красного цвета только для новых строк
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    for row_num in range(start_row, sheet.max_row + 1):
        for col_num in range(1, len(columns) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = red_fill

    wb.save('result.xlsx')
    file_path = 'result.xlsx'
    wb = load_workbook(file_path)

    # Получение активного листа
    sheet = wb.active
    # Установка белого цвета для ячейки (2, 1)
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    sheet.cell(row=2, column=1).fill = white_fill


    # Сохранение изменений
    wb.save('result.xlsx')
    # Загрузка данных из unique_rows_df2 и преобразование в список списков
    unique_rows_2 = unique_rows_df2.values.tolist()

    # Определение столбцов во втором файле Excel
    columns = list(unique_rows_df2.columns)

    # Определение зеленой заливки
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # В цикле по всем строкам Excel документа
    for row_num in range(1, sheet.max_row + 1):
        # Получение строки как списка значений
        row_data = [cell.value for cell in sheet[row_num]]

        # Если эта строка Excel существует в unique_rows_df2
        if row_data in unique_rows_2:
            # Применяем зеленую заливку ко всей строке
            for col_num in range(1, len(columns) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.fill = green_fill

    # Сохранение изменений
    wb.save('result.xlsx')

    append_columns('result.xlsx')
    ps10('result.xlsx')

    # Вызываем функцию с указанными столбцами
    wb = load_workbook('result.xlsx')
    ws = wb.active
    copy_column_style(ws, 'G', ['H', 'I', 'J'])
    wb.save("result.xlsx")