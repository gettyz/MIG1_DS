import openpyxl

# Открываем файл Excel
file_path = 'Машина 1.xlsx'
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# Создаем пустой словарь для хранения уникальных элементов
word_dict = {}

# Проходим по каждой строке (начиная с второй строки)
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=4).value  # Столбец D

    # Проверяем, есть ли слово в списке и нет ли его уже в word_dict
    if cell_value is not None and cell_value not in word_dict:
        word_dict[cell_value] = True  # Используем True для обозначения наличия элемента

# Используем словарь для проверки и записи значений в столбец H
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=6).value  # Столбец F

    # Проверяем, есть ли слово в словаре word_dict
    if cell_value in word_dict:
        worksheet.cell(row=row, column=8, value="Cборка")  # Столбец H
    else:
        worksheet.cell(row=row, column=8, value="Компоненты")  # Столбец H

# Создаем словари Крепеж и Декор
krepezh_set = {"Болт", "Гайка", "Штифт", "Шайба", "Шуруп"}
dekor_set = {"Коврики", "Подушки"}

# Проходим по каждой строке (начиная с второй строки)
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=6).value  # Столбец F

    # Проверяем, к какой категории относится строка
    if cell_value in word_dict:
        category = "Сборка"
    elif cell_value in krepezh_set:
        category = "Крепеж"
    elif cell_value in dekor_set:
        category = "Декор"
    else:
        category = "Детали"

    # Записываем категорию в столбец I
    worksheet.cell(row=row, column=9, value=category)  # Столбец I

# Проходим по каждой строке (начиная с второй строки)
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=5).value  # Столбец Е

    # метод isascii() проверяет, содержатся ли все символы в строке cell_value в стандарте ASCII (American Standard Code for Information Interchange).
    # Стандарт ASCII включает в себя латинские буквы (как заглавные, так и строчные) и цифры, а также некоторые специальные символы.
    if cell_value.isascii() :
        category = "Иностранное"  # Содержит только латиницу и цифры
    elif "ГОСТ" in cell_value:
        category = "РФ по ГОСТ"  # Содержит слово "ГОСТ"
    else:
        category = "РФ"  # Содержит кириллические символы, кроме "ГОСТ"

    # Записываем категорию в столбец J
    worksheet.cell(row=row, column=10, value=category)  # Столбец J


# Сохраняем изменения в файл
workbook.save(file_path)